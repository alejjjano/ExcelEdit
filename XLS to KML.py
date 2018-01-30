import openpyxl
import simplekml
import os


def exceltodict(filename, sheetname):
    """
    filename is a string
    sheet is a string
    returns geodict:
    names as keys strings
    coordinates as values strings
    """

    book = openpyxl.load_workbook(filename)
    sheet = book.get_sheet_by_name(sheetname)

    colm = {}

    j = 1
    while (sheet.cell(row=1, column=j).value) is not None:
        if str(sheet.cell(row=1, column=j).value) in ["Name", "Nombre", "NAME", "NOMBRE", "IE"]:
            colm["NAMES"] = j
        elif str(sheet.cell(row=1, column=j).value) in ["LAT", "Lat"]:
            colm["LATS"] = j
        elif str(sheet.cell(row=1, column=j).value) in ["LONG", "Long"]:
            colm["LONGS"] = j
        elif "NAMES" in colm and "LATS" in colm and "LONGS" in colm:  # If all columns are in colm then stop
            break
        j += 1

    geodict = {}

    i = 2
    while (sheet.cell(row=i, column=1).value) is not None:
        name = str(sheet.cell(row=i, column=colm["NAMES"]).value)
        lat = str(sheet.cell(row=i, column=colm["LATS"]).value)
        long = str(sheet.cell(row=i, column=colm["LONGS"]).value)
        geodict[name] = long + "," + lat
        i += 1

    return geodict


def dicttoKMLfolder(datadict):
    """
    datadict is a dictionary with:
    names as keys, strings
    lat and log as values, strings
    returns KMLfolder with points
    """


def KMLfldtofile(KMLfolder, filename):
    """
    filename is a string
    transforms KMLfolder into a KML file
    and writes it with namef: filename
    """


# Enter main program
# Set files to read and write
filetoread = input("Insert filename to read")
filetoread = filetoread+".xlsx"
filetowrite = filetoread

data = exceltodict(filetoread, "Hoja1")

# Print log of exported data
print("Generate points:")
for key in data.keys():
    print(key, ":", data[key])

folder = dicttoKMLfolder(data)
KMLfldtofile(folder, filetowrite)

# Messages to end program
print("KML wrote with name", filetowrite)
input("Press key to end")
